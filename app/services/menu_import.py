"""Разбор загруженного файла меню (CSV или Excel .xlsx) — чистая логика, без
БД и без FastAPI (см. tests/test_menu_import.py). Вынесено из app/services/menu.py
отдельным модулем, потому что это отдельная ответственность: "превратить
непредсказуемые байты, которые прислал человек, в список строк", а не
"применить эти строки к меню точки".

Файл присылает управляющий чайханы, который не разбирается в технике, и мы
НИКОГДА не видим его заранее — заведение намеренно не делится данными до
подписания (см. задание лупа). Поэтому нельзя рассчитывать ни на формат
(CSV vs Excel), ни на то, что шапка — первая строка (сверху бывает название
заведения, дата, пустые строки, объединённые ячейки), ни на точные
формулировки колонок. Разбор обязан быть терпимым к этому и в случае
проблемы объяснять человеку, что не так, — не трейсбеком, а по-русски.

Различие None и пустого значения в технической карте (composition/allergens/
weight_grams/prep_time_minutes/spiciness) — то же принципиальное требование,
что и в app/ports/menu.py, и оно юридически значимо для аллергенов (spec §5
S2): пустая ячейка = None = "не проверялось" -> ассистент обязан сказать
"уточните на кухне". Явное слово "нет"/"нету"/"отсутствуют" = () = кухня
подтвердила отсутствие аллергенов. Это разные по смыслу утверждения, и разбор
никогда не должен молча превращать одно в другое.
"""

import csv
import io
import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class MenuImportError(Exception):
    """Файл целиком нечитаем: не CSV/XLSX, неизвестная кодировка, нет нужных
    колонок. Сообщение всегда адресовано человеку, который прислал файл
    (app/api/menu.py отдаёт текст этого исключения прямо в 400-ответ) — никаких
    трейсбеков и терминов вроде "InvalidFileException" наружу.

    Не наследуется от MenuServiceError (app/services/menu.py) сознательно: этот
    модуль не знает про остальной сервис меню (БД, стоп-лист, CRUD) и не должен —
    он проверяется юнит-тестами без единого похода в базу."""


# --- Алиасы заголовков --------------------------------------------------------
#
# Колонки ищутся по смыслу заголовка, а не по позиции: порядок колонок в
# реальном файле не гарантирован. Сравнение — по нормализованному тексту
# (см. _normalize_header): регистр, лишние пробелы, перенос строки внутри
# ячейки, буква "ё", знаки препинания — всё это не должно мешать узнать
# колонку. Ключи ниже уже записаны в нормализованном виде.

_HEADER_ALIASES: dict[str, str] = {
    "название": "name",
    "название блюда": "name",
    "наименование": "name",
    "наименование блюда": "name",
    "блюдо": "name",
    "позиция": "name",
    "категория": "category",
    "раздел": "category",
    "группа": "category",
    "тип": "category",
    "цена": "price",
    "стоимость": "price",
    "цена руб": "price",
    "стоимость руб": "price",
    "цена за порцию": "price",
    "состав": "composition",
    "описание": "composition",
    "ингредиенты": "composition",
    "состав блюда": "composition",
    "аллергены": "allergens",
    "вес": "weight_grams",
    "вес г": "weight_grams",
    "выход": "weight_grams",
    "выход г": "weight_grams",
    "граммовка": "weight_grams",
    "порция": "weight_grams",
    "порция г": "weight_grams",
    "время отдачи": "prep_time_minutes",
    "время приготовления": "prep_time_minutes",
    "время": "prep_time_minutes",
    "мин": "prep_time_minutes",
    "острота": "spiciness",
    "острое": "spiciness",
}

# Русское название поля для сообщений об ошибках (человеку, не программисту).
_FIELD_LABELS_RU: dict[str, str] = {
    "name": "название",
    "category": "категория",
    "price": "цена",
    "composition": "состав",
    "allergens": "аллергены",
    "spiciness": "острота",
    "weight_grams": "вес",
    "prep_time_minutes": "время отдачи",
}

# Только однозначные утвердительные слова — "без" отдельно не берём: "без глютена"
# в свободном тексте значило бы другое, а как значение всей ячейки — слишком
# двусмысленно, чтобы автоматически считать это подтверждением "аллергенов нет".
_NO_ALLERGENS_WORDS = frozenset({"нет", "нету", "отсутствуют"})

# Сколько строк сверху просматриваем в поисках шапки. 15 с запасом покрывает
# реальные случаи (название заведения, дата выгрузки, пара пустых строк) и не
# даёт разбору "убежать" по файлу, где шапки нет вовсе.
_MAX_HEADER_SCAN_ROWS = 15


def _normalize_header(cell: object) -> str:
    """Заголовок ячейки -> ключ для сравнения с _HEADER_ALIASES: без регистра,
    без "ё" (сводим к "е"), без переноса строки внутри ячейки (Excel это
    позволяет), без знаков препинания (запятая/точка/скобки трактуются как
    пробел — "цена, руб." и "цена руб" должны совпасть), с одним пробелом
    между словами."""
    if not cell:
        return ""
    text = str(cell).replace("\n", " ").replace("\r", " ")
    text = text.strip().casefold().replace("ё", "е")
    text = re.sub(r"[.,;:()/\\]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _fuzzy_match_field(normalized: str) -> str | None:
    """Точного совпадения с алиасом нет — ищем алиас как отдельное слово (или
    словосочетание) внутри заголовка: реальный заголовок часто длиннее алиаса
    ("цена в рублях за порцию блюда"). Более длинные алиасы проверяются
    первыми, чтобы специфичный "цена за порцию" не проиграл общему "цена"."""
    words = normalized.split()
    word_set = set(words)
    best_field: str | None = None
    best_len = 0
    for alias, field in _HEADER_ALIASES.items():
        alias_words = alias.split()
        matches = alias in word_set if len(alias_words) == 1 else alias in normalized
        if matches and len(alias) > best_len:
            best_field, best_len = field, len(alias)
    return best_field


def _map_known_columns(row: list[str]) -> dict[int, str]:
    """Индекс колонки -> имя поля модели, для одной строки-кандидата в шапку.

    На одно поле в файле нередко претендуют две колонки: «Наименование» и
    «Наименование группы», «Цена» и «Цена со скидкой». Побеждать должна та, что
    названа точнее, а не та, что правее. Раньше выигрывала последняя, и это
    молча ломало весь импорт: блюдо получало имя своей категории, а пустая
    колонка скидки стирала цену и отклоняла каждую строку файла.

    Точным считаем совпадение заголовка с алиасом целиком; «догадка» по
    вхождению слова уступает ему всегда, а при равной точности берём первую
    колонку — в выгрузках главная обычно левее уточняющей.
    """
    column_map: dict[int, str] = {}
    точные: set[str] = set()
    for idx, cell in enumerate(row):
        normalized = _normalize_header(cell)
        if not normalized:
            continue
        точное = _HEADER_ALIASES.get(normalized)
        field = точное or _fuzzy_match_field(normalized)
        if not field:
            continue
        уже = field in column_map.values()
        if уже and (field in точные or точное is None):
            # Поле уже занято — новой колонке нечем перебить занявшую.
            continue
        if уже:
            # Точное совпадение вытесняет прежнюю догадку.
            column_map = {i: f for i, f in column_map.items() if f != field}
        column_map[idx] = field
        if точное:
            точные.add(field)
    return column_map


def _find_header_row(rows: list[list[str]]) -> tuple[int, dict[int, str]]:
    """Найти строку заголовков среди первых _MAX_HEADER_SCAN_ROWS строк файла:
    ту, где угадано больше всего известных колонок. Настоящая шапка почти
    всегда даёт заметно больше совпадений, чем случайное слово в тексте выше
    неё, поэтому берём строку с максимумом; при равном счёте — более позднюю
    (ближе к данным, значит вероятнее настоящая шапка, а не случайное
    совпадение в преамбуле)."""
    best_idx = -1
    best_map: dict[int, str] = {}
    best_score = 0
    for idx, row in enumerate(rows[:_MAX_HEADER_SCAN_ROWS]):
        column_map = _map_known_columns(row)
        if column_map and len(column_map) >= best_score:
            best_score = len(column_map)
            best_idx = idx
            best_map = column_map
    if best_idx < 0:
        raise MenuImportError(_header_not_found_message(rows))
    return best_idx, best_map


def _header_not_found_message(rows: list[list[str]]) -> str:
    sample = [
        cell.strip()
        for row in rows[:_MAX_HEADER_SCAN_ROWS]
        for cell in row
        if cell and cell.strip()
    ][:10]
    hint = ", ".join(f"«{cell}»" for cell in sample) if sample else "в файле нет непустых ячеек"
    return (
        f"Не нашли строку с заголовками колонок в первых {_MAX_HEADER_SCAN_ROWS} "
        "строках файла. Нужны как минимум колонки «название» и «цена» (можно другими "
        "словами, например «наименование» и «стоимость»). "
        f"Первые непустые ячейки файла, которые мы увидели: {hint}."
    )


def _describe_header(header_cells: list[str], column_map: dict[int, str]) -> str:
    """Что мы поняли из строки заголовков — для сообщения об отсутствующей
    обязательной колонке: управляющий должен увидеть, что мы уже распознали,
    и что осталось нераспознанным, чтобы понять, какую колонку переименовать."""
    recognized = [
        f"«{header_cells[idx].strip()}» — это «{_FIELD_LABELS_RU[field]}»"
        for idx, field in sorted(column_map.items())
        if idx < len(header_cells) and header_cells[idx].strip()
    ]
    unrecognized = [
        f"«{cell.strip()}»"
        for idx, cell in enumerate(header_cells)
        if cell and cell.strip() and idx not in column_map
    ]
    parts = []
    if recognized:
        parts.append("Мы распознали: " + "; ".join(recognized) + ".")
    if unrecognized:
        parts.append("Остальные колонки файла: " + ", ".join(unrecognized) + ".")
    return " ".join(parts)


def _missing_column_error(
    field: str, header_cells: list[str], column_map: dict[int, str]
) -> MenuImportError:
    label = _FIELD_LABELS_RU[field]
    message = f"В файле нет обязательной колонки «{label}»."
    description = _describe_header(header_cells, column_map)
    if description:
        message += " " + description
    message += f" Переименуйте нужную колонку в «{label}» и загрузите файл заново."
    return MenuImportError(message)


# --- Декодирование и разбор CSV -----------------------------------------------


def decode_menu_csv(raw: bytes) -> str:
    """Декодировать CSV/текстовый файл. Приоритет UTF-8: если сотрудник выгрузил
    файл сам, это чаще всего он. CP1251 — второй в очереди осознанно: это
    реальный случай экспорта из Excel под русской Windows (разделитель там
    обычно ";"), а не гипотетический."""
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise MenuImportError("Не удалось определить кодировку файла (ожидается UTF-8 или CP1251)")


def _sniff_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=";,").delimiter
    except csv.Error:
        # Короткий/однострочный файл — Sniffer не справляется. По умолчанию ";":
        # это разделитель Excel под русской локалью, самый вероятный реальный случай.
        return ";" if sample.count(";") >= sample.count(",") else ","


def _rows_from_csv(text: str) -> list[list[str]]:
    delimiter = _sniff_delimiter(text[:2048])
    return list(csv.reader(io.StringIO(text), delimiter=delimiter))


# --- Разбор Excel (.xlsx) ------------------------------------------------------
#
# Формат определяется по сигнатуре байтов, а не по расширению файла: человек,
# не разбирающийся в технике, может прислать файл с любым или вовсе без
# расширения. .xlsx — это ZIP-архив (сигнатура "PK\x03\x04"), старый бинарный
# .xls — формат OLE2 (сигнатура ниже). Всё остальное пробуем как CSV/текст.

_XLSX_MAGIC = b"PK\x03\x04"
_XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _detect_kind(raw: bytes) -> str:
    if raw.startswith(_XLSX_MAGIC):
        return "xlsx"
    if raw.startswith(_XLS_MAGIC):
        return "xls"
    return "csv"


def _cell_to_text(value: object) -> str:
    """Ячейка Excel -> текст в духе CSV-ячейки, чтобы дальше работал один и тот
    же разбор строк независимо от формата файла. Числа — самый частый случай
    (цена, вес, острота почти всегда приходят как число, не текст): str(float)
    в Python — кратчайшая строка, которая парсится обратно в то же число, тех
    же проблем с 0.1+0.2 здесь не возникает, потому что мы только читаем
    исходное значение ячейки, а не считаем поверх него."""
    if value is None:
        return ""
    if isinstance(value, bool):  # bool — подкласс int, проверяем раньше него
        return "да" if value else "нет"
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, datetime):
        # Сравниваем с полуночью явно: time(0, 0) — истинный объект, поэтому
        # проверка «if value.time()» всегда срабатывала, и дата без времени
        # печаталась как «01.02.2026 00:00».
        без_времени = value.time() == time(0, 0)
        return value.strftime("%d.%m.%Y") if без_времени else value.strftime("%d.%m.%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value).strip()


def _rows_from_xlsx(raw: bytes) -> list[list[str]]:
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except (InvalidFileException, KeyError, zipfile.BadZipFile, OSError) as exc:
        raise MenuImportError(
            "Не удалось открыть файл как Excel (.xlsx) — похоже, он повреждён "
            "или на самом деле не является файлом .xlsx."
        ) from exc
    try:
        if not workbook.worksheets:
            raise MenuImportError("В файле Excel нет ни одного листа")
        # Первый лист — реальные меню почти всегда однолистовые, а называть
        # листы управляющий может как угодно, выбирать по имени не на чем.
        sheet = workbook.worksheets[0]
        # Читаем в try не для порядка: в режиме read_only openpyxl разбирает XML
        # листа лениво, уже во время обхода строк. Поэтому недокачанный или
        # побитый файл — обычное дело, когда его копируют с флешки или шлют
        # мессенджером — падает НЕ на открытии, а здесь. Без этого перехвата
        # управляющий получал 500 и трейсбек вместо объяснения.
        return [
            [_cell_to_text(cell) for cell in raw_row]
            for raw_row in sheet.iter_rows(values_only=True)
        ]
    except MenuImportError:
        raise
    except Exception as exc:  # noqa: BLE001 — любая поломка разбора XML листа
        raise MenuImportError(
            "Файл Excel открылся, но прочитать лист не удалось — скорее всего, "
            "он повреждён или скачался не полностью. Откройте его в Excel, "
            "сохраните заново (Файл → Сохранить как → Книга Excel) и пришлите ещё раз."
        ) from exc
    finally:
        workbook.close()


# --- Разбор одной строки данных (общий для CSV и Excel) -----------------------


_ЧИСЛО_В_ЯЧЕЙКЕ = re.compile(r"-?\d+")

# Что в ячейке цены пишут кроме самой цены: знак рубля, слово, пробел-разделитель
# тысяч (в выгрузках он бывает и обычным, и неразрывным).
# Порядок важен: длинные варианты первыми, иначе «руб» съест начало «рублей»
# и оставит «лей», а цена перестанет разбираться.
_МУСОР_В_ЦЕНЕ = re.compile(r"(?i)рублей|рубля|руб\.?|р\.|[₽$€\s  ]")


def _очистить_цену(raw: str) -> str:
    """«1 200 руб» и «520,50 ₽» — это цена, а не поломка файла.

    Цена — обязательное поле: если её не разобрать, теряется всё блюдо, а не
    одна ячейка. Выгрузки из учётных систем регулярно приходят с разделителем
    тысяч и валютой прямо в ячейке, и отказываться от такого файла значит
    отказаться от половины меню заведения.
    """
    return _МУСОР_В_ЦЕНЕ.sub("", raw).replace(",", ".")


def _parse_optional_int(
    raw: str,
    *,
    field_name: str,
    errors: list[str],
    min_value: int | None = None,
    max_value: int | None = None,
) -> int | None:
    """Разобрать необязательное целое поле.

    Единицы измерения в самой ячейке — норма, а не поломка: «250 г», «15 мин»,
    «0,5 кг» пишут ровно так же часто, как выносят в заголовок. Достаём число и
    работаем дальше.

    Граница проведена так: **не поняли — молчим и не мешаем блюду; поняли, и
    значение заведомо неверное — говорим.**

    Совсем нечитаемая ячейка оставляет поле пустым, но строку НЕ бракует.
    Раньше любая такая мелочь роняла блюдо целиком: «Лагман, 450 ₽, 250 г» не
    попадал в меню из-за буквы «г» в весе. Вес и время подачи — справочные, а
    название и цена у строки уже есть; терять из-за них блюдо, которое официант
    потом не найдёт, — плохой размен.

    Выход за диапазон — другое дело: число мы прочитали, и «острота 5» при шкале
    0–3 значит, что в заведении своя шкала. Это стоит показать управляющему в
    предпросмотре, а не проглотить молча.
    """
    if not raw:
        return None
    try:
        value = int(raw)
    except ValueError:
        совпадение = _ЧИСЛО_В_ЯЧЕЙКЕ.search(raw)
        if совпадение is None:
            return None
        try:
            value = int(совпадение.group())
        except ValueError:
            return None
    too_small = min_value is not None and value < min_value
    too_large = max_value is not None and value > max_value
    if too_small or too_large:
        errors.append(f"«{field_name}» вне диапазона {min_value}..{max_value}: {value}")
        return None
    return value


@dataclass(frozen=True)
class MenuCsvRow:
    """Одна разобранная строка файла. errors непусто — строка не будет применена.
    Имя оставлено историческим ("Csv"), хотя источником теперь может быть и
    Excel — переименовывать смысла нет, тип уже везде на виду (app/api/menu.py,
    app/schemas/menu.py, scripts/seed_venue.py)."""

    line_number: int
    name: str
    category: str | None
    price: Decimal | None
    composition: str | None
    allergens: tuple[str, ...] | None
    spiciness: int | None
    weight_grams: int | None
    prep_time_minutes: int | None
    errors: tuple[str, ...] = ()

    @property
    def is_valid(self) -> bool:
        return not self.errors


def _parse_row(line_number: int, values: dict[str, str]) -> MenuCsvRow:
    errors: list[str] = []

    name = values.get("name", "")
    if not name:
        errors.append("не заполнено название")

    category = values.get("category") or None

    price_raw = values.get("price", "")
    price: Decimal | None = None
    if not price_raw:
        errors.append("не заполнена цена")
    else:
        try:
            price = Decimal(_очистить_цену(price_raw))
        except InvalidOperation:
            errors.append(f"не удалось разобрать цену: {price_raw!r}")
        else:
            if price < 0:
                errors.append(f"цена не может быть отрицательной: {price_raw!r}")
                price = None

    composition = values.get("composition") or None

    allergens_raw = values.get("allergens", "")
    allergens: tuple[str, ...] | None
    if not allergens_raw:
        allergens = None
    elif allergens_raw.strip().casefold() in _NO_ALLERGENS_WORDS:
        allergens = ()
    else:
        allergens = tuple(a.strip() for a in allergens_raw.split(",") if a.strip())

    spiciness = _parse_optional_int(
        values.get("spiciness", ""), field_name="острота", errors=errors, min_value=0, max_value=3
    )
    weight_grams = _parse_optional_int(
        values.get("weight_grams", ""), field_name="вес", errors=errors, min_value=0
    )
    prep_time_minutes = _parse_optional_int(
        values.get("prep_time_minutes", ""), field_name="время отдачи", errors=errors, min_value=0
    )

    return MenuCsvRow(
        line_number=line_number,
        name=name,
        category=category,
        price=price,
        composition=composition,
        allergens=allergens,
        spiciness=spiciness,
        weight_grams=weight_grams,
        prep_time_minutes=prep_time_minutes,
        errors=tuple(errors),
    )


# --- Точка входа ----------------------------------------------------------------


def parse_menu_file(raw: bytes, *, filename: str | None = None) -> list[MenuCsvRow]:
    """Разобрать загруженный файл (CSV или Excel .xlsx) в список строк. Формат
    определяется по содержимому, не по имени файла — `filename` используется
    только для более понятного текста ошибки. Чистая функция — без БД, без
    FastAPI, легко проверяется юнит-тестами (см. tests/test_menu_import.py)."""
    kind = _detect_kind(raw)
    if kind == "xls":
        hint = f" («{filename}»)" if filename else ""
        raise MenuImportError(
            f"Файл{hint} сохранён в старом формате Excel 97-2003 (.xls) — мы такой не "
            "читаем. Откройте его в Excel и пересохраните: «Файл» → «Сохранить как» → "
            "«Книга Excel (.xlsx)», затем загрузите получившийся файл."
        )

    rows = _rows_from_xlsx(raw) if kind == "xlsx" else _rows_from_csv(decode_menu_csv(raw))
    if not rows or not any(any(cell.strip() for cell in row) for row in rows):
        raise MenuImportError("Файл пуст")

    header_idx, column_map = _find_header_row(rows)
    header_cells = rows[header_idx]
    present_fields = set(column_map.values())
    if "name" not in present_fields:
        raise _missing_column_error("name", header_cells, column_map)
    if "price" not in present_fields:
        raise _missing_column_error("price", header_cells, column_map)

    parsed: list[MenuCsvRow] = []
    for line_number, raw_row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if not any(cell.strip() for cell in raw_row):
            continue  # полностью пустая строка — не ошибка данных, просто пропуск
        values = {
            field_name: (raw_row[idx].strip() if idx < len(raw_row) else "")
            for idx, field_name in column_map.items()
        }
        parsed.append(_parse_row(line_number, values))
    return parsed
